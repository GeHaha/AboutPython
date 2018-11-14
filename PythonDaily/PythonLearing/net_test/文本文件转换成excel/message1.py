# -*- coding: utf-8 -*-
"""
Created on Fri Jun 29 14:00:12 2018

@author: Administrator
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment,Font,colors
import re

wb = Workbook()

dest_file = ('H:\Python\python learnig\test\文本文件转换成excel\student.xlsx')

ws = wb.create_sheet(title = 'student')

alignment = Alignment(horizontal='center', vertical='center')
font = Font(bold=True,color=colors.BLUE)

col = 'ABCDE'

ws[col[0]+'1'] = '#'
ws[col[1]+'1'] = '姓名'
ws[col[2]+'1'] = '数学'
ws[col[3]+'1'] = '英语'
ws[col[4]+'1'] = '语文'

for x in col:
    ws[x+'1'].alignment = alignment
    ws[x+'1'].font = font

data = []

p = re.compile(':\[')

source_file = open('H:\Python\python learnig\test\文本文件转换成excel\student.txt','r')

for line in source_file:
    if not line.startswith('{') and not line.startswith('}'):
        line = line.strip('\n], ')
        line = p.sub(',',line)
        data.append(line.split(','))

#print data
source_file.close()

for i in range(len(data)):
    for j in range(5):
        ws[col[j]+str(i+2)] = data[i][j].strip('"')
        ws[col[j]+str(i+2)].alignment = alignment

wb.save(dest_file)