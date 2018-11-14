# -*- coding: utf-8 -*-
"""
Created on Fri Jun 29 09:39:46 2018

@author: Administrator
"""
#将txt文件的内容读取出来写入到Excel文件中
"""
import xlrd    #写入文件
import xlwt    #打开Excel文件
import os


workbook = xlwt.Workbook(encoding = 'utf-8')
os.mknod('H:\Python\python learnig\test\文本文件转换成excel\student.xls')

#worksheet = workbook.add_sheet('sheet1')


txtFileName = ('student.txt')

excelFileName = ('H:\Python\python learnig\test\文本文件转换成excel\student.xls')
#新建一个excel文件
"""
"""
file = xlwt.Workbook(encoding = 'utf-8',style_compression = 0 )

fp = open(txtFileName)

for line in fp.readlines():
    worksheet.write(0,0,label = line)
workbook.save('student.xls')
"""
""" 
from openpyxl import workbook

wb = workbook()
ws = wb.active()
ws1 = wb.creat_sheet(0)
ws.title = ("New Title")
"""

import json
import xlwt
from collections import OrderedDict
 
 
def run_1():
 
    with open('student.txt','r') as f:
        content = f.read()
 
    #转化为json，注意转化后的dict的元素位置可能和转化前可能不一样，因此需要ordereddict
    #loads()方法把str对象反序列化为json对象，自定义解码器为ordereddict
    d = json.loads(content,object_pairs_hook=OrderedDict)
    print(d)
    #初始化xls文件
    file = xlwt.Workbook()
    #添加sheet,工作表，名字为test
    table = file.add_sheet('test')
    for row ,i in enumerate(d):   #读取所有字典，row为序号，i为字典关键字key
        table.write(row,0,i)    #写入（行号，列号，key)
        for col,j in enumerate(d[i]):   #col为序号，j为value,有多个，需要迭代
            table.write(row,col+1,j)
 
    file.save('H:\Python\python learnig\test\文本文件转换成excel\student.xls')









































 