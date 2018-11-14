# -*- coding: utf-8 -*-
"""
Created on Fri Jun 29 10:58:01 2018

@author: Administrator
"""

#encoding: utf8 
from openpyxl import Workbook


#创建一个工作薄
#使用openpyxl没有必要先在系统中新建一个.xlsx，我们需要做的只需要引Workbook这个类，接着开始调用它。
wb = Workbook()

#一个工作簿(workbook)在创建的时候同时至少也新建了一张工作表(worksheet)。你可以通过openpyxl.workbook.Workbook.active()调用得到正在运行的工作表。
ws = wb.active
"""
注意：该函数调用工作表的索引(_active_sheet_index)，默认是0。除非你修改了这个值，否则你使用该函数一直是在对第一张工作表进行操作。
"""

#使用openpyxl.workbook.Workbook.create_sheet()新建一张表
ws1 = wb.create_sheet()   #默认插在工作簿末尾
ws2 = wb.create_sheet(0)  #插入在工作簿的第一个位置


#在创建工作表的时候系统自动命名。他们按照序列依次命名 (Sheet, Sheet1, Sheet2, ...)。你可以通过调用下面的属性修改工作表的名称：
ws.title = "New Title"   #系统第一张表默认的名称Sheet


#标签栏的背景色默认为白色。你可以通过提供一个RRGGBB颜色码改变标签栏的字体
#ws.sheet_properties.tabColor = "1072BA"
ws.sheet_properties.tabColor = "ff0033" #红色


#一旦你获取工作表的名字，你可以通过workbook的key或者openpyxl.workbook.Workbook.get_sheet_by_name() 方法得到该工作表
ws3 = wb["New Title"]
ws4 = wb.get_sheet_by_name("New Title")
print(ws is ws3 is ws4)


#你可以通过openpyxl.workbook.Workbook.get_sheet_names() 方法得到工作簿的所有工作表。
print(wb.get_sheet_names())


#你也可以循环得到所有的工作表
for sheet in wb:
    print(sheet.title)


#单元格可以直接根据他们的索引直接获得
c = ws['A4']
print(type(c))
print(c)


#通过上述的语句，将返回在A4处的单元格，如果不存在将在A4新建一个。 单元格的
ws['A4'] = 4


#还提供 openpyxl.worksheet.Worksheet.cell() 方法获取单元格
c = ws.cell("A4")
print(c)


#也可以根据行列值获取单元格
d = ws.cell(row = 4, column = 2)
print(d)
print("d:",type(d))

"""
注意：当一个工作表被创建是，其中不包含单元格。只有当单元格被获取是才被创建。这种方式我们不会创建我们从不会使用的单元格，从而减少了内存消耗。
"""


#警告：由于上述特性，你如果遍历了单元格而非想要使用它们也将会在内存当中创建。比如下面：
for i in range(1,101):
    for j in range(1,101):
        ws.cell(row = i, column = j)

"""
上述代码将会在内存中创建100*100个单元格。当然，这里也有方法来清理这些不想要的单元格，在后续我们将会介绍。
"""       


#使用切片获取多个单元格
cell_range = ws['A1':'C2']


#使用openpyxl.worksheet.Worksheet.iter_rows() 方法获得多个单元格
tuple(ws.iter_rows('A1:C2'))
print(tuple(ws.iter_rows('A1:C2')))

list(ws.iter_rows('A1:C2'))
print(list(ws.iter_rows('A1:C2')))

for row in ws.iter_rows('A1:C2'):
    for cell in row:
        print(cell)
        
#如果你需要迭代文件中所有的行或者列，你可以使用,openpyxl.worksheet.Worksheet.rows()
ws['C9'] = 'hello world'


#一旦我们有一个openpyxl.cell.Cell，我们可以直接为该单元格赋值
c = ws['C8']
c.value = "nihao"
print(c.value)
print(ws['C8'].value)

d.value = 3.14
print(d.value)


#你也可以使用Python中的其他类型和格式
#wb = Workbook(guess_types=True)
c.value = '12%'
print(c.value)
import datetime
d.value = datetime.datetime.now()
print (d.value)
c.value = '31.50'
print(c.value)


#保存工作簿最简单最安全的方式是使用openpyxl.workbook.Workbook的openpyxl.workbook.Workbook.save() 方法
#wb.save("TonyDemoExcle.xlsx")

"""
特别警告：这个操作将会在没有认识提示的情况下用现在写的内容，覆盖掉原文件中的所有内容
"""

#你也可以 as_template=True，将文件保存称为一个模板
#wb = load_workbook('TonyDemoExcle1.xlsx')
#wb.save('TonyDemoExcle1.xltx', as_template=True)


#和写入文件的方式相同，你可以引入openpyxl.load_workbook()来打开一个已经存在的工作簿

#from openpyxl import load_workbook
#wb2 = load_workbook('test.xlsx')
#print(wb2.get_sheet_names())